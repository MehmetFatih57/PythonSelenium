from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains  # For click method fallback
from selenium.webdriver.support.ui import Select  # For dropdown handling
from PIL import Image  # For screenshot functionality (optional)
import time  # For time-based waits (consider using explicit waits instead)
import openpyxl
import os


class JavascriptExecutor:
    pass


class ReusableMethods:
    def bekle(self, saniye):
        try:
            time.sleep(saniye)
        except KeyboardInterrupt:
            raise RuntimeError("Kesme oluştu")

    # Alert handling (if needed):
    def alert_accept(self):
        driver = self.get_driver()  # Assuming you have a way to get the driver instance
        driver.switch_to.alert().accept()

    @staticmethod
    def alert_dismiss(self):
        driver = self.get_driver()
        driver.switch_to.alert().dismiss()

    def alert_text(self):
        driver = self.get_driver()
        return driver.switch_to.alert().text

    def alert_prompt(self, text):
        driver = self.get_driver()
        driver.switch_to.alert().send_keys(text)

    # Dropdowns (using Select):
    def ddm_visible_text(self, element, text):
        select = Select(element)
        select.select_by_visible_text(text)

    def ddm_index(self, element, index):
        select = Select(element)
        select.select_by_index(index)

    def ddm_value(self, element, value):
        select = Select(element)
        select.select_by_value(value)

    # Window switching:
    def switch_to_window(self, index):
        driver = self.get_driver()
        handles = driver.window_handles
        driver.switch_to.window(handles[index])

    # Explicit wait methods:
    def visible_wait(self, element, seconds):
        wait = WebDriverWait(self.get_driver(), seconds)
        wait.until(EC.visibility_of(element))

    def visible_element_locator_wait(self, locator, seconds):
        wait = WebDriverWait(self.get_driver(), seconds)
        return wait.until(EC.visibility_of_element_located(locator))

    def alert_wait(self, seconds):
        wait = WebDriverWait(self.get_driver(), seconds)
        wait.until(EC.alert_is_present())

    # Screenshot (using PIL, adapt as needed):
    def full_page_screenshot(self, name):
        driver = self.get_driver()
        page_screenshot = driver.get_screenshot_as_png()
        image = Image.open(page_screenshot)
        timestamp = time.strftime("%H_%M_%S_%d%m%Y")  # Adapt timestamp format if needed
        image.save(f"Screenshots/{name}{timestamp}.png")  # Replace with your desired path

    # Click method with fallback for non-clickable elements:
    def click(self, element):
        try:
            element.click()
        except (Exception,):
            actions = ActionChains(self.get_driver())
            actions.move_to_element(element).perform()
            actions.click(element).perform()

    #   JavaScriptexecution methods:
    def scroll_to_element(self, element):
        driver = self.get_driver()
        js = JavascriptExecutor(driver)
        js.execute_script("arguments[0].scrollIntoView(true);", element)

    def scroll_to_bottom(self):
        driver = self.get_driver()
        js = JavascriptExecutor(driver)
        js.executeScript("window.scrollTo(0, document.body.scrollHeight)")

    def scroll_to_top(self):
        driver = self.get_driver()
        js = JavascriptExecutor(driver)
        js.executeScript("window.scrollTo(0, -document.body.scrollHeight)")

    def send_keys_js(self, element, text):
        driver = self.get_driver()
        js = JavascriptExecutor(driver)
        js.executeScript("arguments[0].value='" + text + "'", element)

    def jsClick(self, element):
        driver = self.get_driver
        js = JavascriptExecutor(driver)
        js.execute_script("arguments[0].click();", element)

    # Veya ortam değişkeni kullanın (Çözüm 2)
    # DATA_FILE_PATH = os.environ["DATA_FILE_PATH"]

    data_file_path ="../data/kullanici_bilgileri.xlsx",
    sheet_name = "Sayfa1"


    @staticmethod
    def getData(data_file_path=None, sheet_name=None):
        """
        Kullanıcı bilgilerini belirtilen Excel dosyasından okur.

        Args:
            data_file_path (str, optional): Excel dosyasının yolu. Varsayılan olarak tanımlı değilse hata verir.
            sheet_name (str, optional): Kullanıcı bilgilerinin bulunduğu sheet adı. Varsayılan olarak "Sayfa1" ve "Sheet1" aranır.

        Returns:
            list: Kullanıcı adı ve şifrelerden oluşan liste. Her eleman bir tuple (username, password) şeklindedir.

        Raises:
            FileNotFoundError: Dosya bulunamadığında hata fırlatır.
        """

        if not data_file_path:
            raise ValueError("Excel dosya yolu gereklidir.")

        try:
            excelFile = openpyxl.load_workbook(data_file_path)
        except FileNotFoundError:
            print(f"Dosya bulunamadı: {data_file_path}")
            return []

        # Sheet adını kontrol et
        selectedSheet = None
        if sheet_name:
            selectedSheet = excelFile.get(sheet_name)
        else:
            # Varsayılan sheet adlarını dene
            for name in ["Sayfa1", "Sheet1"]:
                if name in excelFile.sheetnames:
                    selectedSheet = excelFile[name]
                    break

        if not selectedSheet:
            print(f"Sheet bulunamadı: {sheet_name if sheet_name else 'Varsayılan sheet adları'}")
            return []

        rows = selectedSheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            data.append((username, password))

        return data
