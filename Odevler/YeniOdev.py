from telnetlib import EC
from time import sleep
from utilitiessss import Constant as c

from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from utilitiessss import Driver, ConfigReader
from pages import SaucePage
import openpyxl
import pytest


class TestSauce:
    def setup_method(self):
        self.driver = Driver.get_driver(True)
        self.driver.get(ConfigReader.read_config("url"))
        self.page = SaucePage.SaucePage(self.driver)

    def teardown_method(self):
        self.driver.quit()

    @staticmethod
    def getData():
        excelFile = openpyxl.load_workbook("../data/kullanici_bilgileri.xlsx")
        selectedSheet = excelFile["Sayfa1"]
        rows = selectedSheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            data.append((username, password))

        return data

    message = "Epic sadface: Sorry, this user has been locked out."

    @pytest.mark.parametrize("username,password", getData())
    def test_odev(self, username, password):
        pass
        self.page.user_name.send_keys(username)
        sleep(2)
        self.page.password.send_keys(password)
        sleep(2)
        self.page.login_button.click()
        sleep(2)
        if (c.url_inventory == self.driver.current_url):
            print("Girdin")
        elif (c.error_locked_user == self.message):
            print("Kilidi acmak icin parmaginizi okutunuz")
        else:
            "Giremedi ki , Giremedi ki "
