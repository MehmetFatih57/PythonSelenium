import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from utilitiessss import Driver, ConfigReader
from pages import SaucePage

class Test_SauceDemo:
    def setup_method(self):
        self.driver = Driver.get_driver()
        self.driver.get(ConfigReader.read_config("url"))
        self.page = SaucePage.SaucePage(self.driver)

    def teardown_method(self):
        self.driver.quit()


    def getData():
        excelFile = openpyxl.load_workbook("../data/invalidLogin.xlsx")
        selectedSheet = excelFile["Sheet1"]
        rows = selectedSheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            data.append((username, password))

        return data

    @pytest.mark.parametrize("username,password", getData())
    def test_invalid_login(self, username, password):
        pass
        self.page.user_name.send_keys(username)
        self.page.password.send_keys(password)
        self.page.login_button.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    @pytest.mark.skip
    def test_valid_login(self):
        self.page.username.send_keys("standard_user")
        self.page.password.send_keys("secret_sauce")
        self.page.loginButton.click()
        appLogo = self.driver.find_element(By.CLASS_NAME, "app_logo")
        assert appLogo.text == "Swag Labs"